import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class ExcelUtil:

    @staticmethod
    def read_excel(path):
        return pd.read_excel(path)

    @staticmethod
    def write_excel(df, path, sheet_name="Sheet1"):
        df.to_excel(path, index=False, sheet_name=sheet_name)

    @staticmethod
    def format_header(path):
        wb = load_workbook(path)
        ws = wb.active

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid"
            )

        wb.save(path)

    @staticmethod
    def display_excel(df):
        # lấy cột Mã + Tên
        data = df.iloc[:, 2:4]

        return data